# -*- coding: utf-8 -*-
"""
@author: Shweta Nagre
"""

# choosing possible primers
import pandas as pd
import configparser
import sys
import openpyxl
name1 = sys.argv[1]

config = configparser.ConfigParser()

#read elimination conditions from config file
config.read_file(open(r'config.txt'))
t1 = config.get('My Section', 'temperature(>)',fallback='No Value Entered')
t2 = config.get('My Section', 'temperature(<)',fallback='No Value Entered')
gc1 = config.get('My Section', 'gc(>)',fallback='No Value Entered')
gc2 = config.get('My Section', 'gc(<)',fallback='No Value Entered')
t1=float(t1)
t2=float(t2)
gc1=float(gc1)
gc2=float(gc2)

wrkbk = openpyxl.load_workbook(name1+"_MIPs.xlsx")
sh = wrkbk.active

elim=[]
good=[]
#check for temperature and gc content limits, eliminate the row if the condition(temp or gc) isn't met for both arms
for i in range(1, sh.max_row+1):
    row=[]
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
        row.append(cell_obj.value)
    if i == 1:
        continue
    else:
        if float(row[4]) > t1 or float(row[4]) < t2:
            elim.append(row)
        elif float(row[5]) < gc2 or float(row[5]) > gc1:
            elim.append(row)
        elif float(row[8]) > t1 or float(row[8]) < t2:
            elim.append(row)
        elif float(row[9]) < gc2 or float(row[9]) > gc1:
            elim.append(row)
        elif row[6] == 'yes' or row[10] == 'yes':
            elim.append(row)
        else:
            good.append(row)
        
        
df = pd.DataFrame(elim, columns = ["Def Line","Main Sequence","Target region","Ligation Arm","TM 1","GC content 1","Continuous stretch","Extension Arm","TM 2","GC content 2","Continuous stretch 2","Reverse Strand Target region","Ligation Arm(Rev)","TM Rev","GC content Rev","Continuous stretch Rev","Extension Arm Rev","TM 2 Rev","GC content 2 Rev","Continuous stretch 2 Rev","Extension Start","Extension End","Ligation Start","Ligation End","Organism"])
df2 = pd.DataFrame(good, columns = ["Def Line","Main Sequence","Target region","Ligation Arm","TM 1","GC content 1","Continuous stretch","Extension Arm","TM 2","GC content 2","Continuous stretch 2","Reverse Strand Target region","Ligation Arm(Rev)","TM Rev","GC content Rev","Continuous stretch Rev","Extension Arm Rev","TM 2 Rev","GC content 2 Rev","Continuous stretch 2 Rev","Extension Start","Extension End","Ligation Start","Ligation End","Organism"])

df.to_excel("Eliminated MIPs.xlsx", index=False)
df2.to_excel("Passable MIPs.xlsx", index=False)
